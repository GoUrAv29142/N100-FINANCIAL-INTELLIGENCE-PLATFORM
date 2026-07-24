import sqlite3
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from src.logger import get_logger

log = get_logger(__name__)

DB_PATH = Path("db") / "nifty100.db"

RADAR_AXES = [
    ("return_on_equity_pct", "ROE"),
    ("return_on_capital_employed_pct", "ROCE"),
    ("net_profit_margin_pct", "NPM"),
    ("debt_to_equity", "D/E (inv)"),
    ("free_cash_flow_cr", "FCF"),
    ("pat_cagr_5yr", "PAT CAGR 5yr"),
    ("revenue_cagr_5yr", "Revenue CAGR 5yr"),
    ("composite_quality_score", "Composite Score"),
]

COL_TO_METRIC = {
    "return_on_equity_pct": "roe",
    "return_on_capital_employed_pct": "roce",
    "net_profit_margin_pct": "npm",
    "debt_to_equity": "de",
    "free_cash_flow_cr": "fcf",
    "pat_cagr_5yr": "pat_cagr_5yr",
    "revenue_cagr_5yr": "revenue_cagr_5yr",
    "eps_cagr_5yr": "eps_cagr_5yr",
    "interest_coverage": "interest_coverage",
    "asset_turnover": "asset_turnover",
}


def load_data_for_charts() -> pd.DataFrame:
    conn = sqlite3.connect(DB_PATH)
    df = pd.read_sql_query("""
        SELECT fr.*, pg.peer_group_name
        FROM financial_ratios fr
        LEFT JOIN peer_groups pg ON fr.company_id = pg.company_id
        WHERE (fr.company_id, fr.year) IN (
            SELECT company_id, MAX(year) FROM financial_ratios GROUP BY company_id
        )
    """, conn)
    conn.close()
    return df


def _normalize_for_radar(series: pd.Series, invert=False) -> pd.Series:
    s = series.astype(float)
    valid = s.dropna()
    if len(valid) < 2 or valid.min() == valid.max():
        return s.apply(lambda x: 50.0 if pd.notna(x) else None)
    scaled = (s - valid.min()) / (valid.max() - valid.min()) * 100
    if invert:
        scaled = 100 - scaled
    return scaled


def generate_radar_chart(company_id: str, df: pd.DataFrame, output_dir: Path):
    company_row = df[df["company_id"] == company_id]
    if company_row.empty:
        log.warning("No data for %s, skipping radar chart", company_id)
        return
    company_row = company_row.iloc[0]

    peer_group = company_row.get("peer_group_name")
    has_peer_group = pd.notna(peer_group)

    comparison_pool = df[df["peer_group_name"] == peer_group] if has_peer_group else df
    comparison_label = peer_group if has_peer_group else "Nifty 100 Average"

    labels = [label for _, label in RADAR_AXES]
    company_values = []
    peer_avg_values = []

    for col, _ in RADAR_AXES:
        invert = (col == "debt_to_equity")
        normed_pool = _normalize_for_radar(df[col], invert=invert)

        cval = normed_pool.loc[company_row.name] if company_row.name in normed_pool.index else None
        company_values.append(cval if pd.notna(cval) else 0)

        pool_idx = comparison_pool.index
        pool_vals = normed_pool.loc[normed_pool.index.intersection(pool_idx)]
        peer_avg_values.append(pool_vals.mean() if len(pool_vals) > 0 else 0)

    num_vars = len(labels)
    angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()
    company_values += company_values[:1]
    peer_avg_values += peer_avg_values[:1]
    angles += angles[:1]

    fig, ax = plt.subplots(figsize=(7, 7), subplot_kw=dict(polar=True))
    ax.plot(angles, company_values, linewidth=2, linestyle="solid", label=company_id, color="#2E86AB")
    ax.fill(angles, company_values, alpha=0.25, color="#2E86AB")
    ax.plot(angles, peer_avg_values, linewidth=2, linestyle="dashed", label=comparison_label, color="#A23B72")

    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_ylim(0, 100)
    ax.set_title(f"{company_id} vs {comparison_label}", fontsize=13, pad=20)
    ax.legend(loc="upper right", bbox_to_anchor=(1.3, 1.1), fontsize=9)

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / f"{company_id}_radar.png"
    plt.tight_layout()
    plt.savefig(out_path, dpi=100, bbox_inches="tight")
    plt.close(fig)


def generate_all_radar_charts():
    df = load_data_for_charts()
    output_dir = Path("reports") / "radar_charts"
    for company_id in df["company_id"].unique():
        generate_radar_chart(company_id, df, output_dir)
    log.info("Generated %d radar charts in %s", df["company_id"].nunique(), output_dir)


def generate_peer_comparison_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font

    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    GOLD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    HEADER_FONT = Font(bold=True)

    conn = sqlite3.connect(DB_PATH)
    ratios = pd.read_sql_query("""
        SELECT fr.*, pg.peer_group_name, pg.is_benchmark, c.company_name
        FROM financial_ratios fr
        JOIN peer_groups pg ON fr.company_id = pg.company_id
        LEFT JOIN companies c ON fr.company_id = c.id
        WHERE (fr.company_id, fr.year) IN (
            SELECT company_id, MAX(year) FROM financial_ratios GROUP BY company_id
        )
    """, conn)
    percentiles = pd.read_sql_query("SELECT * FROM peer_percentiles", conn)
    conn.close()

    metric_cols = list(COL_TO_METRIC.keys())

    pct_lookup = {
        (row["company_id"], row["peer_group_name"], row["metric"]): row["percentile_rank"]
        for _, row in percentiles.iterrows()
    }

    wb = Workbook()
    wb.remove(wb.active)

    for group_name, group_df in ratios.groupby("peer_group_name"):
        sheet_name = str(group_name)[:31]
        ws = wb.create_sheet(sheet_name)

        header = ["company_id", "company_name"] + metric_cols
        ws.append(header)
        for cell in ws[1]:
            cell.font = HEADER_FONT

        for _, row in group_df.iterrows():
            ws.append([row.get(c) for c in header])
            r = ws.max_row

            if row.get("is_benchmark") == 1:
                for c in range(1, len(header) + 1):
                    ws.cell(row=r, column=c).fill = GOLD
                continue

            for i, col in enumerate(metric_cols, start=3):
                metric_key = COL_TO_METRIC[col]
                p = pct_lookup.get((row["company_id"], group_name, metric_key))
                if p is not None:
                    fill = GREEN if p >= 0.75 else (YELLOW if p >= 0.25 else RED)
                    ws.cell(row=r, column=i).fill = fill

        median_row = ["MEDIAN", ""] + [
            group_df[c].median() if c in group_df.columns else None for c in metric_cols
        ]
        ws.append(median_row)
        for cell in ws[ws.max_row]:
            cell.font = HEADER_FONT

    Path("output").mkdir(exist_ok=True)
    out_path = Path("output") / "peer_comparison.xlsx"
    wb.save(out_path)
    log.info("Generated %s with %d sheets", out_path, len(wb.sheetnames))
    return out_path


if __name__ == "__main__":
    import logging
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

    print("Generating radar charts...")
    generate_all_radar_charts()

    print("Generating peer_comparison.xlsx...")
    generate_peer_comparison_xlsx()

    print("Done.")
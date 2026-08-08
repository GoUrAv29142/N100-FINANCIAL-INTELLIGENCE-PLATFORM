"""
src/screener/engine.py

Day 15: Filter Engine Core.

Loads config/screener_config.yaml and applies threshold filters to the
financial_ratios table (latest year per company only - screening on
stale historical data doesn't make business sense). Supports all 15
filterable metrics from the Sprint 3 spec, with two special-case rules:

  - D/E filters are automatically skipped for companies in the
    Financials broad_sector, since high leverage is structurally
    normal for banks/NBFCs/insurers (same rule as Sprint 2 Day 13).
  - ICR filters treat icr_label == 'Debt Free' as passing any minimum
    threshold (debt-free companies have no interest expense, so their
    true interest coverage is effectively infinite).
"""

import sqlite3
from pathlib import Path
from typing import Optional

import pandas as pd
import yaml

from nifty100_dashboard.src.logger import get_logger

log = get_logger(__name__)

DB_PATH = Path("db") / "nifty100.db"
CONFIG_PATH = Path("config") / "screener_config.yaml"


def load_config() -> dict:
    """Load screener_config.yaml."""
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def load_latest_year_ratios() -> pd.DataFrame:
    """
    Load financial_ratios joined with sectors, restricted to each
    company's latest available year only.
    """
    conn = sqlite3.connect(DB_PATH)

    query = """
        SELECT fr.*, s.broad_sector
        FROM financial_ratios fr
        LEFT JOIN sectors s ON fr.company_id = s.company_id
        WHERE (fr.company_id, fr.year) IN (
            SELECT company_id, MAX(year) FROM financial_ratios GROUP BY company_id
        )
    """
    df = pd.read_sql_query(query, conn)
    conn.close()

    df["is_financials"] = df["broad_sector"] == "Financials"
    return df


def apply_filter(df: pd.DataFrame, filter_key: str, threshold, filter_defs: dict) -> pd.DataFrame:
    """
    Apply a single named filter (e.g. 'roe_min') to the DataFrame.
    Returns the filtered DataFrame.
    """
    if filter_key not in filter_defs:
        log.warning("Unknown filter key '%s' - skipping", filter_key)
        return df

    spec = filter_defs[filter_key]
    column = spec["column"]
    direction = spec["direction"]

    if column not in df.columns:
        log.warning("Filter column '%s' not found in DataFrame - skipping filter '%s'", column, filter_key)
        return df

    working = df.copy()

    # ICR special case: Debt Free companies always pass a minimum ICR threshold
    if spec.get("debt_free_always_passes") and "icr_label" in working.columns:
        debt_free_mask = working["icr_label"] == "Debt Free"
    else:
        debt_free_mask = pd.Series([False] * len(working), index=working.index)

    # D/E special case: skip the filter entirely for Financials companies
    if spec.get("skip_for_financials") and "is_financials" in working.columns:
        financials_mask = working["is_financials"] == True
    else:
        financials_mask = pd.Series([False] * len(working), index=working.index)

    if direction == "min":
        passes = working[column] >= threshold
    elif direction == "max":
        passes = working[column] <= threshold
    else:
        raise ValueError(f"Unknown filter direction: {direction}")

    # A row passes if it meets the threshold, OR it's exempted by a special case
    final_mask = passes | debt_free_mask | financials_mask

    return working[final_mask]


def apply_filters(df: pd.DataFrame, filters: dict, filter_defs: dict) -> pd.DataFrame:
    """Apply a dict of {filter_key: threshold} filters in sequence."""
    result = df.copy()
    for filter_key, threshold in filters.items():
        if filter_key == "de_equals":
            # Special exact-match filter for Debt-Free Blue Chip preset
            if "debt_to_equity" in result.columns:
                result = result[result["debt_to_equity"].notna() & (result["debt_to_equity"] == threshold)]
            continue
        result = apply_filter(result, filter_key, threshold, filter_defs)
    return result


def run_custom_screen(filters: dict) -> pd.DataFrame:
    """
    Run a custom filter set against the latest-year financial_ratios data.
    filters: dict like {"roe_min": 15, "de_max": 1.0}
    Returns sorted DataFrame with composite_quality_score, descending.
    """
    config = load_config()
    df = load_latest_year_ratios()
    filtered = apply_filters(df, filters, config["filters"])
    return filtered.sort_values("composite_quality_score", ascending=False)
  
def run_preset(preset_key: str) -> pd.DataFrame:
    """
    Run one of the 6 named preset screeners defined in screener_config.yaml.
    Handles the two presets with special logic beyond simple threshold
    filters (Dividend Champion's payout ceiling, Debt-Free Blue Chip's
    exact D/E=0 match, Turnaround Watch's YoY D/E decline check).
    """
    config = load_config()
    presets = config["presets"]

    if preset_key not in presets:
        raise ValueError(f"Unknown preset: {preset_key}. Valid: {list(presets.keys())}")

    preset = presets[preset_key]
    df = load_latest_year_ratios()

    filtered = apply_filters(df, preset.get("filters", {}), config["filters"])

    # Dividend Champion: additional payout ceiling (not a min/max column filter,
    # it's a maximum on dividend_payout_ratio_pct specifically)
    if "dividend_payout_max" in preset:
        max_payout = preset["dividend_payout_max"]
        filtered = filtered[
            filtered["dividend_payout_ratio_pct"].isna()
            | (filtered["dividend_payout_ratio_pct"] <= max_payout)
        ]

    # Turnaround Watch: D/E must be declining year-over-year (latest vs prior year)
    if preset.get("de_declining_yoy"):
        filtered = _filter_declining_de_yoy(filtered)

    rank_by = preset.get("rank_by", "composite_quality_score")
    rank_desc = preset.get("rank_desc", True)

    if rank_by in filtered.columns:
        filtered = filtered.sort_values(rank_by, ascending=not rank_desc)

    return filtered


def _filter_declining_de_yoy(df: pd.DataFrame) -> pd.DataFrame:
    """
    Keep only companies whose D/E in their latest year is lower than
    their D/E in the prior available year (Turnaround Watch preset).
    """
    conn = sqlite3.connect(DB_PATH)
    keep_ids = []

    for cid in df["company_id"]:
        history = pd.read_sql_query(
            "SELECT year, debt_to_equity FROM financial_ratios WHERE company_id = ? ORDER BY year",
            conn, params=(cid,)
        )
        if len(history) < 2:
            continue
        latest_de = history.iloc[-1]["debt_to_equity"]
        prior_de = history.iloc[-2]["debt_to_equity"]
        if pd.notna(latest_de) and pd.notna(prior_de) and latest_de < prior_de:
            keep_ids.append(cid)

    conn.close()
    return df[df["company_id"].isin(keep_ids)]
def _winsorize_and_score(series: pd.Series, invert: bool = False) -> pd.Series:
    """
    Winsorize a series at P10/P90 and scale to 0-100.
    Same logic as Sprint 2's composite_quality_score, reused here for
    the Screener's separate, spec-defined composite score.
    """
    s = series.copy().astype(float)
    valid = s.dropna()
    if len(valid) < 2:
        return pd.Series([None] * len(s), index=s.index)

    p10 = valid.quantile(0.10)
    p90 = valid.quantile(0.90)
    clipped = s.clip(lower=p10, upper=p90)

    if p90 == p10:
        return pd.Series([50.0 if pd.notna(v) else None for v in s], index=s.index)

    score = (clipped - p10) / (p90 - p10) * 100
    if invert:
        score = 100 - score

    return score.where(s.notna(), None)


def compute_screener_composite_score(df: pd.DataFrame, sector_relative: bool = True) -> pd.DataFrame:
    """
    Day 17: Screener composite quality score (0-100).

    35% Profitability (ROE 15% + ROCE 10% + NPM 10%)
    30% Cash Quality (FCF CAGR 15% + CFO/PAT ratio 10% + FCF positive flag 5%)
    20% Growth (Revenue CAGR 10% + PAT CAGR 10%)
    15% Leverage (D/E score 10% + ICR score 5%)

    Each input metric is winsorized at P10/P90 and scaled 0-100 before
    weighting. If sector_relative=True, winsorization is computed
    separately within each broad_sector rather than across all 92
    companies, per the spec's "sector-relative composite score"
    requirement (Day 17).

    Distinct from Sprint 2's composite_quality_score (different formula,
    different weights) - stored as screener_composite_score to avoid
    confusion between the two scores.
    """
    df = df.copy()

    # FCF positive flag: 100 if free_cash_flow_cr > 0, else 0, None if missing
    df["_fcf_positive_flag"] = df["free_cash_flow_cr"].apply(
        lambda x: 100.0 if pd.notna(x) and x > 0 else (0.0 if pd.notna(x) else None)
    )
    # Outlier guard: ROE values beyond a sane range are near-zero-
    # denominator artifacts (documented in Sprint 2's ratio_edge_cases.log
    # - e.g. HAL, BEL, INDIGO), not genuine performance. Excluding them
    # from the scoring INPUT (not just capping via winsorization) prevents
    # them from dominating sector-relative rankings, since winsorization
    # alone still preserves their rank position even after capping the
    # displayed magnitude.
    df["_roe_for_scoring"] = df["return_on_equity_pct"].where(
        df["return_on_equity_pct"].abs() <= 500, None
    )
    df["_roce_for_scoring"] = df["return_on_capital_employed_pct"].where(
        df["return_on_capital_employed_pct"].abs() <= 500, None
    )

    metric_cols = {
        "roe": ("_roe_for_scoring", False),
        "roce": ("_roce_for_scoring", False),
        "npm": ("net_profit_margin_pct", False),
        "fcf_cagr": ("fcf_cagr_5yr", False),
        "cfo_pat": ("cfo_pat_ratio", False),
        "fcf_flag": ("_fcf_positive_flag", False),
        "revenue_cagr": ("revenue_cagr_5yr", False),
        "pat_cagr": ("pat_cagr_5yr", False),
        "de": ("debt_to_equity", True),   # inverted: lower D/E = higher score
        "icr": ("interest_coverage", False),
    }

    scores = {}

    if sector_relative and "broad_sector" in df.columns:
        for key, (col, invert) in metric_cols.items():
            scored = pd.Series([None] * len(df), index=df.index, dtype=object)
            for sector, group in df.groupby("broad_sector", dropna=False):
                sector_score = _winsorize_and_score(group[col], invert=invert)
                scored.loc[group.index] = sector_score
            scores[key] = scored
    else:
        for key, (col, invert) in metric_cols.items():
            scores[key] = _winsorize_and_score(df[col], invert=invert)

    # ICR "Debt Free" companies get max ICR score (they have no interest
    # burden at all, which is the best possible leverage outcome)
    if "icr_label" in df.columns:
        debt_free_mask = df["icr_label"] == "Debt Free"
        scores["icr"] = scores["icr"].where(~debt_free_mask, 100.0)

    weights = {
        "roe": 0.15, "roce": 0.10, "npm": 0.10,          # Profitability 35%
        "fcf_cagr": 0.15, "cfo_pat": 0.10, "fcf_flag": 0.05,  # Cash Quality 30%
        "revenue_cagr": 0.10, "pat_cagr": 0.10,          # Growth 20%
        "de": 0.10, "icr": 0.05,                          # Leverage 15%
    }

    def _combine(idx):
        parts = [(scores[k].get(idx), w) for k, w in weights.items()]
        valid_parts = [(s, w) for s, w in parts if s is not None and pd.notna(s)]
        if not valid_parts:
            return None
        total_w = sum(w for _, w in valid_parts)
        weighted = sum(s * w for s, w in valid_parts)
        return round(weighted / total_w, 2)

    df["screener_composite_score"] = [_combine(idx) for idx in df.index]

    return df.drop(columns=["_fcf_positive_flag"])


def generate_screener_output():
    """
    Day 17: output/screener_output.xlsx - one sheet per preset (6 total),
    20 KPI columns, sorted by composite score descending, colour-coded
    (green = meets preset threshold, red = fails it).
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    HEADER_FONT = Font(bold=True)

    display_cols = [
        "company_id", "return_on_equity_pct", "return_on_capital_employed_pct",
        "net_profit_margin_pct", "debt_to_equity", "interest_coverage",
        "asset_turnover", "free_cash_flow_cr", "fcf_cagr_5yr", "cfo_pat_ratio",
        "revenue_cagr_5yr", "pat_cagr_5yr", "eps_cagr_5yr", "pe_ratio",
        "pb_ratio", "dividend_yield_pct", "market_cap_crore", "sales",
        "net_profit", "screener_composite_score",
    ]

    config = load_config()
    wb = Workbook()
    wb.remove(wb.active)

    for preset_key in PRESET_KEYS:
        preset = config["presets"][preset_key]
        result = run_preset(preset_key)
        result = compute_screener_composite_score(result, sector_relative=True)
        result = result.sort_values("screener_composite_score", ascending=False)

        sheet_name = preset["name"][:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)

        cols_present = [c for c in display_cols if c in result.columns]
        ws.append(cols_present)
        for cell in ws[1]:
            cell.font = HEADER_FONT

        preset_filters = preset.get("filters", {})
        filter_defs = config["filters"]

        for _, row in result.iterrows():
            ws.append([row.get(c) for c in cols_present])
            r = ws.max_row
            for filter_key, threshold in preset_filters.items():
                if filter_key not in filter_defs:
                    continue
                col_name = filter_defs[filter_key]["column"]
                direction = filter_defs[filter_key]["direction"]
                if col_name not in cols_present:
                    continue
                col_idx = cols_present.index(col_name) + 1
                val = row.get(col_name)
                if val is None or pd.isna(val):
                    continue
                passed = (val >= threshold) if direction == "min" else (val <= threshold)
                ws.cell(row=r, column=col_idx).fill = GREEN if passed else RED

        for i, col in enumerate(cols_present, 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

    Path("output").mkdir(exist_ok=True)
    out_path = Path("output") / "screener_output.xlsx"
    wb.save(out_path)
    log.info("Generated %s", out_path)
    return out_path







PRESET_KEYS = [
    "quality_compounder",
    "value_pick",
    "growth_accelerator",
    "dividend_champion",
    "debt_free_blue_chip",
    "turnaround_watch",
]

if __name__ == "__main__":
    print("=== Preset Screener Sanity Check ===\n")
    for key in PRESET_KEYS:
        result = run_preset(key)
        status = "OK" if 5 <= len(result) <= 50 else "CHECK (outside 5-50 range)"
        print(f"{key}: {len(result)} companies [{status}]")
        if len(result) > 0:
            print(result[["company_id", "composite_quality_score"]].head(5).to_string(index=False))
        print()

    print("Generating output/screener_output.xlsx...")
    generate_screener_output()
    print("Done.")